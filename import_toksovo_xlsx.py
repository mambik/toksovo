#!/usr/bin/env python3
import json
import os
import re
import shutil
import ssl
import unicodedata
from pathlib import Path
from urllib.parse import quote, urlparse
from urllib.request import Request, urlopen

import openpyxl

WORKBOOK_PATH = Path("/Users/sergey/Downloads/toksovo.xlsx")
ASSET_ROOT = Path("assets/imported")
YANDEX_API = "https://cloud-api.yandex.net/v1/disk/public/resources"
USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"
UNVERIFIED_CONTEXT = ssl._create_unverified_context()
MAX_PHOTOS = 1
USE_REMOTE_PHOTOS = os.environ.get("TOKSOVO_REMOTE_PHOTOS") == "1"

CYRILLIC_TO_LATIN = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "e",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "h",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "shch",
    "ъ": "",
    "ы": "y",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
}

THEME_BY_CATEGORY = {
    "Достопримечательности": "attractions",
    "Скрытые локации": "hidden",
    "Дети": "children",
    "Дети и взрослые": "children",
    "Танцы": "children",
    "Лыжная школа": "sport",
    "Спорт для взрослых": "sport",
    "Взрослые": "sport",
    "Студенты": "sport",
}


def clean_text(value):
    value = "" if value is None else str(value)
    value = value.replace("\xa0", " ").replace("\u200b", " ")
    value = re.sub(r"\s+", " ", value).strip()
    if value in {"", "#ERROR!"}:
        return ""
    return value


def first_line(value):
    value = clean_text(value)
    if not value:
        return ""
    return value.splitlines()[0].strip()


def extract_first_url(value):
    text = clean_text(value)
    if not text:
        return ""
    if text.lower() in {"нет информации", "дубль", "больше не работает"}:
        return ""
    match = re.search(r"https?://[^\s]+", text, re.I)
    if match:
        return match.group(0).rstrip(".,);")
    if re.match(r"^[\w.-]+\.[a-z]{2,}(?:/.*)?$", text, re.I):
        return f"https://{text}"
    return ""


def normalize_multi_text(*values):
    parts = []
    for value in values:
        text = clean_text(value)
        if text and text not in parts:
            parts.append(text)
    return "\n".join(parts)


def transliterate(value):
    normalized = unicodedata.normalize("NFKD", clean_text(value).lower())
    pieces = []
    for char in normalized:
        if char in CYRILLIC_TO_LATIN:
            pieces.append(CYRILLIC_TO_LATIN[char])
        elif char.isascii():
            pieces.append(char)
    text = "".join(pieces)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return text or "place"


def unique_slug(name, seen):
    base = transliterate(name)
    slug = base
    counter = 2
    while slug in seen:
        slug = f"{base}-{counter}"
        counter += 1
    seen.add(slug)
    return slug


def theme_for_category(category):
    return THEME_BY_CATEGORY.get(clean_text(category), "hidden")


def load_rows():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in values):
            continue
        rows.append({
            "id": ws.cell(r, 1).value,
            "name": ws.cell(r, 2).value,
            "category": ws.cell(r, 3).value,
            "subcategory": ws.cell(r, 4).value,
            "age": ws.cell(r, 5).value,
            "description": ws.cell(r, 6).value,
            "address": ws.cell(r, 7).value,
            "district": ws.cell(r, 8).value,
            "format": ws.cell(r, 9).value,
            "price": ws.cell(r, 10).value,
            "phone": ws.cell(r, 11).value,
            "website": ws.cell(r, 12).value,
            "lat": ws.cell(r, 13).value,
            "lng": ws.cell(r, 14).value,
            "rating": ws.cell(r, 15).value,
            "source": ws.cell(r, 16).value,
            "status": ws.cell(r, 17).value,
            "photo_storage": ws.cell(r, 18).value,
            "note_2": ws.cell(r, 19).value,
            "note_3": ws.cell(r, 20).value,
        })
    return rows


def fetch_json(url):
    req = Request(url, headers={"Accept": "application/json", "User-Agent": USER_AGENT})
    with urlopen(req, timeout=20, context=UNVERIFIED_CONTEXT) as resp:
        return json.load(resp)


def fetch_public_folder(public_url):
    api_url = (
        f"{YANDEX_API}?public_key={quote(public_url, safe='')}"
        "&limit=1000&preview_size=XL"
    )
    data = fetch_json(api_url)
    embedded = data.get("_embedded") or {}
    items = embedded.get("items") or []
    if not items:
        return []

    images = []
    for item in items:
        if item.get("type") != "file":
            continue
        if not str(item.get("mime_type") or "").startswith("image/"):
            continue
        source_url = item.get("preview") or ""
        for size in item.get("sizes") or []:
            if size.get("name") == "ORIGINAL" and size.get("url"):
                source_url = size["url"]
                break
        if not source_url:
            continue
        images.append({
            "name": item.get("name") or "photo",
            "url": source_url,
        })
    def sort_key(item):
        parts = re.split(r"(\d+)", item["name"].lower())
        return [int(part) if part.isdigit() else part for part in parts]

    images.sort(key=sort_key)
    return images[:MAX_PHOTOS]


def download_file(url, dest):
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and dest.stat().st_size > 0:
        return
    req = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(req, timeout=30, context=UNVERIFIED_CONTEXT) as resp:
        with open(dest, "wb") as fh:
            shutil.copyfileobj(resp, fh)


def file_ext_from_name(name, fallback=".jpg"):
    suffix = Path(name).suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        return suffix
    return fallback


def build_place(row, seen_slugs, folder_cache):
    name = clean_text(row["name"])
    category = clean_text(row["category"])
    category_id = theme_for_category(category)
    slug = unique_slug(name, seen_slugs)
    website = extract_first_url(row["website"])
    notes = normalize_multi_text(
        row["note_2"],
        row["note_3"],
    )

    place = {
        "id": slug,
        "slug": slug,
        "name": name,
        "category": category,
        "categoryId": category_id,
        "subcategory": clean_text(row["subcategory"]),
        "description": clean_text(row["description"]),
        "address": clean_text(row["address"]),
        "district": clean_text(row["district"]),
        "phone": clean_text(row["phone"]),
        "website": website,
        "price": clean_text(row["price"]),
        "age": clean_text(row["age"]),
        "format": clean_text(row["format"]),
        "rating": clean_text(row["rating"]),
        "source": clean_text(row["source"]),
        "status": clean_text(row["status"]),
        "lat": row["lat"] if row["lat"] not in (None, "") else None,
        "lng": row["lng"] if row["lng"] not in (None, "") else None,
    }

    if notes:
        place["notes"] = notes

    return place


def write_output(path, places):
    text = "const PLACES = " + json.dumps(places, ensure_ascii=False, indent=2) + ";\n"
    path.write_text(text, encoding="utf-8")


def main():
    rows = load_rows()
    seen_slugs = set()
    folder_cache = {}
    places = []
    total = len(rows)
    for index, row in enumerate(rows, start=1):
        name = clean_text(row["name"])
        if index == 1 or index % 10 == 0 or index == total:
            print(f"[{index}/{total}] {name}")
        places.append(build_place(row, seen_slugs, folder_cache))

    write_output(Path("data.js"), places)
    Path("data.json").write_text(json.dumps(places, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"Imported {len(places)} places from {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
